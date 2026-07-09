"""excel_summary_writer 单测：写出回读对账/计数口径/失败折叠/公式注入防护/mock 水印。"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from tools_impl.excel_summary_writer.adapter import run

_CASES = [
    {
        "case_id": "case_001",
        "status": "success",
        "params": {"altitude_m": 1000.0, "mach": 0.3, "power_kw": 800.0},
        "outputs": {"shaft_power_kw": 700.0, "fuel_flow_kgps": 0.06, "egt_c": 450.0},
        "error_message": None,
        "mock": True,
    },
    {
        "case_id": "case_002",
        "status": "failed",
        "params": {"altitude_m": 16000.0, "mach": 0.5, "power_kw": 900.0},
        "outputs": None,
        "error_message": "超出 mock 包线",
        "mock": True,
    },
]


def test_write_then_read_back_rows_and_counts(tmp_path: Path) -> None:
    out = tmp_path / "result_summary.xlsx"
    result = run({"cases": _CASES, "output_path": str(out)})

    assert result["status"] == "success"
    assert result["file_path"] == str(out)
    assert result["ok_count"] == 1
    assert result["failed_count"] == 1
    assert out.is_file()

    ws = openpyxl.load_workbook(out)["result_summary"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        "case_id", "altitude_m", "mach", "power_kw",
        "shaft_power_kw", "fuel_flow_kgps", "egt_c", "status", "error_message", "mock",
    )
    assert len(rows) == 1 + len(_CASES)
    assert rows[1][0] == "case_001" and rows[1][7] == "success"
    assert rows[2][0] == "case_002" and rows[2][7] == "failed"
    assert rows[2][8] == "超出 mock 包线"
    # 失败 case 的输出列如实为空，绝不填伪造数值
    assert rows[2][4] is None and rows[2][5] is None and rows[2][6] is None
    # mock 末列逐行如实转写
    assert rows[1][9] is True and rows[2][9] is True


def test_mock_column_defaults_to_false_when_absent(tmp_path: Path) -> None:
    """case 记录缺 mock 键 → 末列按 false 写出（不无中生有替数据源声明 mock）。"""
    out = tmp_path / "nomock.xlsx"
    cases = [{"case_id": "case_001", "status": "success", "params": {"altitude_m": 1.0}}]
    result = run({"cases": cases, "output_path": str(out)})
    assert result["status"] == "success"
    ws = openpyxl.load_workbook(out)["result_summary"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][-1] == "mock"
    assert rows[1][-1] is False


def test_formula_injection_forced_inert_text(tmp_path: Path) -> None:
    """P1（CWE-1236）：`=` 开头的未受信字符串必须以惰性文本落盘（data_type='s'），
    值原样保留、绝不成为活公式。若去掉 adapter 的强制 data_type 行，本测试必红
    （openpyxl 默认把 '=...' 存为 data_type='f' 公式）。
    """
    evil_case_id = '=HYPERLINK("http://evil","x")'
    evil_error = "=1+1"
    out = tmp_path / "inject.xlsx"
    result = run({
        "cases": [{
            "case_id": evil_case_id,
            "status": "failed",
            "params": {"altitude_m": 1000.0},
            "outputs": None,
            "error_message": evil_error,
        }],
        "output_path": str(out),
    })
    assert result["status"] == "success"

    # data_only=False：读原始单元格（含公式定义），检真实存储形态
    ws = openpyxl.load_workbook(out, data_only=False)["result_summary"]
    header = [c.value for c in ws[1]]
    id_cell = ws.cell(row=2, column=header.index("case_id") + 1)
    err_cell = ws.cell(row=2, column=header.index("error_message") + 1)

    assert id_cell.data_type == "s", f"case_id 单元格必须是惰性文本，实为 {id_cell.data_type!r}"
    assert err_cell.data_type == "s", f"error_message 单元格必须是惰性文本，实为 {err_cell.data_type!r}"
    assert id_cell.value == evil_case_id, "内容必须原样保留（不删改，只灭活）"
    assert err_cell.value == evil_error


def test_notice_writes_declaration_sheet_after_data_sheet(tmp_path: Path) -> None:
    """mock 水印（docs/03 §3 第五落点）：notice 有值 → 数据 sheet 保持第一，
    其后追加「声明」sheet 含声明文本。"""
    out = tmp_path / "notice.xlsx"
    notice = "MOCK 声明：本文件结果无任何工程意义。"
    result = run({"cases": _CASES, "output_path": str(out), "notice": notice})
    assert result["status"] == "success"

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["result_summary", "声明"], "数据 sheet 必须保持第一，声明 sheet 随后"
    assert wb["声明"]["A1"].value == notice


def test_no_notice_no_declaration_sheet(tmp_path: Path) -> None:
    out = tmp_path / "plain.xlsx"
    result = run({"cases": _CASES, "output_path": str(out)})
    assert result["status"] == "success"
    assert openpyxl.load_workbook(out).sheetnames == ["result_summary"]


def test_empty_cases_writes_header_only(tmp_path: Path) -> None:
    out = tmp_path / "empty.xlsx"
    result = run({"cases": [], "output_path": str(out)})
    assert result["status"] == "success"
    assert result["ok_count"] == 0 and result["failed_count"] == 0
    ws = openpyxl.load_workbook(out).active
    assert len(list(ws.iter_rows(values_only=True))) == 1


def test_unwritable_output_path_folds_to_failed(tmp_path: Path) -> None:
    out = tmp_path / "no_such_dir" / "x.xlsx"  # 目录不存在
    result = run({"cases": _CASES, "output_path": str(out)})
    assert result["status"] == "failed"
    assert result["error_message"]
    assert not out.exists()
