"""excel_summary_writer 适配器（mock=false，真实写文件）。

只忠实写出调用方给的 case 结果，不做任何计算/判断/过滤：
- 列序固定：case_id → 参数列（按 _PARAM_ORDER 中实际出现者）→ 输出列
  （按 _OUTPUT_ORDER 中实际出现者）→ status → error_message → mock。
- ok_count/failed_count 用 `== "success"` / `== "failed"` 判定（安全 gate
  一律显式比较，不用 truthiness）。
- 适配器绝不抛裸异常（docs/03）：写盘失败折叠为 {"status":"failed", ...}。

安全（CWE-1236 公式注入，M3 反审 P1）：case_id/error_message 等字符串来自
用户上传表（未受信），openpyxl 对 `=` 开头字符串默认存为活公式（data_type='f'）
——本适配器对**所有**写入的 str 值强制 `cell.data_type = "s"`（白名单式：
全部字符串一律惰性文本，不做 `=`/`+`/`@` 前缀黑名单判断），上游内容
永远不会成为可执行公式。

mock 水印（docs/03 §3 第五落点）：数据 sheet 末列 `mock` 逐行如实转写
（取 case 记录的 mock 布尔，缺省 false）；调用方传 `notice` 时在数据 sheet
（保持第一）之后追加名为「声明」的 sheet 写入该文本——表格产物脱离任务
上下文单独传播时，声明随行。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

_PARAM_ORDER = ("altitude_m", "mach", "power_kw", "bleed_flow_kgps")
_OUTPUT_ORDER = ("shaft_power_kw", "fuel_flow_kgps", "egt_c")

_NOTICE_SHEET = "声明"


def _write_row(ws: Any, row_idx: int, values: list[Any]) -> None:
    """逐格写入；str 值一律强制惰性文本（data_type='s'），杜绝公式注入。"""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if isinstance(value, str):
            cell.data_type = "s"


def run(payload: dict[str, Any], context: dict[str, Any] | None = None) -> dict[str, Any]:
    del context  # V0.1 未用
    cases: list[dict[str, Any]] = payload["cases"]
    output_path = Path(payload["output_path"])
    notice = payload.get("notice")

    param_cols = [
        p for p in _PARAM_ORDER
        if any(p in (c.get("params") or {}) for c in cases)
    ]
    output_cols = [
        o for o in _OUTPUT_ORDER
        if any(o in (c.get("outputs") or {}) for c in cases)
    ]

    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "result_summary"
        _write_row(ws, 1, ["case_id", *param_cols, *output_cols, "status", "error_message", "mock"])

        ok_count = 0
        failed_count = 0
        for row_idx, case in enumerate(cases, start=2):
            params = case.get("params") or {}
            outputs = case.get("outputs") or {}
            status = case.get("status")
            if status == "success":
                ok_count += 1
            else:
                failed_count += 1
            _write_row(ws, row_idx, [
                case.get("case_id"),
                *[params.get(p) for p in param_cols],
                *[outputs.get(o) for o in output_cols],
                status,
                case.get("error_message"),
                bool(case.get("mock", False)),
            ])

        if notice is not None:
            ws_notice = wb.create_sheet(_NOTICE_SHEET)  # 追加在数据 sheet 之后，数据 sheet 保持第一
            _write_row(ws_notice, 1, [notice])

        wb.save(output_path)
    except Exception as exc:  # noqa: BLE001 - 契约要求绝不裸抛
        return {"status": "failed", "error_message": f"汇总表写出失败：{exc.__class__.__name__}: {exc}"}

    return {
        "status": "success",
        "file_path": str(output_path),
        "ok_count": ok_count,
        "failed_count": failed_count,
    }
