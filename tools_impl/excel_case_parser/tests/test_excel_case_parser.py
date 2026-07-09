"""excel_case_parser 单测：正常解析 + 反例必咬（坏文件/空表/缺列/行级错误）。"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from tools_impl.excel_case_parser.adapter import run

HEADER = ["case_id", "altitude_m", "mach", "power_kw"]


def _write_xlsx(path: Path, header: list, rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_parse_valid_table(tmp_path: Path) -> None:
    p = _write_xlsx(tmp_path / "ok.xlsx", HEADER, [
        ["case_001", 1000, 0.3, 800],
        ["case_002", 5000, 0.5, 1200],
    ])
    result = run({"file_path": str(p)})
    assert result["status"] == "success"
    assert result["errors"] == []
    assert [c["case_id"] for c in result["cases"]] == ["case_001", "case_002"]
    assert result["cases"][0]["params"] == {"altitude_m": 1000.0, "mach": 0.3, "power_kw": 800.0}


def test_parse_optional_bleed_column(tmp_path: Path) -> None:
    p = _write_xlsx(tmp_path / "bleed.xlsx", HEADER + ["bleed_flow_kgps"], [
        ["case_001", 1000, 0.3, 800, 1.5],
        ["case_002", 2000, 0.4, 900, None],  # 可选列缺值=不提供该参数，不是行错误
    ])
    result = run({"file_path": str(p)})
    assert result["status"] == "success"
    assert result["cases"][0]["params"]["bleed_flow_kgps"] == 1.5
    assert "bleed_flow_kgps" not in result["cases"][1]["params"]
    assert result["errors"] == []


def test_row_level_errors_skip_but_keep_parsing(tmp_path: Path) -> None:
    """缺值/非数值/超范围/重复 case_id/缺 case_id：进 errors 跳过，不炸整表。"""
    p = _write_xlsx(tmp_path / "dirty.xlsx", HEADER, [
        ["case_001", 1000, 0.3, 800],       # 合法
        ["case_002", None, 0.4, 900],       # altitude 缺值
        ["case_003", 2000, "abc", 900],     # mach 非数值
        ["case_004", 2000, 9.0, 900],       # mach 超范围（>3）
        ["case_001", 3000, 0.5, 700],       # case_id 重复
        [None, 3000, 0.5, 700],             # case_id 缺失
        ["case_005", 99999, 0.5, 700],      # altitude 超合理范围（>30000）
        ["case_006", 16000, 0.5, 700],      # >15000 但 ≤30000：合法！包线判定归 mock 工具
    ])
    result = run({"file_path": str(p)})
    assert result["status"] == "success"
    assert [c["case_id"] for c in result["cases"]] == ["case_001", "case_006"]
    assert len(result["errors"]) == 6
    rows_with_error = [e["row"] for e in result["errors"]]
    assert rows_with_error == [3, 4, 5, 6, 7, 8]  # Excel 行号（表头=1，数据从 2 起）


def test_missing_required_column_is_parser_failure(tmp_path: Path) -> None:
    p = _write_xlsx(tmp_path / "nocol.xlsx", ["case_id", "altitude_m", "mach"], [
        ["case_001", 1000, 0.3],
    ])
    result = run({"file_path": str(p)})
    assert result["status"] == "failed"
    assert "power_kw" in result["error_message"]


def test_nonexistent_file_is_parser_failure(tmp_path: Path) -> None:
    result = run({"file_path": str(tmp_path / "ghost.xlsx")})
    assert result["status"] == "failed"
    assert "不存在" in result["error_message"]


def test_non_xlsx_file_is_parser_failure(tmp_path: Path) -> None:
    p = tmp_path / "fake.xlsx"
    p.write_text("这不是一个 xlsx 文件", encoding="utf-8")
    result = run({"file_path": str(p)})
    assert result["status"] == "failed"


def test_header_only_table_is_parser_failure(tmp_path: Path) -> None:
    p = _write_xlsx(tmp_path / "empty.xlsx", HEADER, [])
    result = run({"file_path": str(p)})
    assert result["status"] == "failed"
    assert "零有效 case" in result["error_message"]


def test_all_rows_invalid_is_parser_failure(tmp_path: Path) -> None:
    p = _write_xlsx(tmp_path / "allbad.xlsx", HEADER, [
        ["case_001", None, 0.3, 800],
        ["case_002", 1000, "x", 800],
    ])
    result = run({"file_path": str(p)})
    assert result["status"] == "failed"
    assert "零有效 case" in result["error_message"]
    assert "2" in result["error_message"]  # 行级错误计数如实上报


def test_sheet_name_selection_and_missing_sheet(tmp_path: Path) -> None:
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "wrong_sheet"
    ws1.append(["unrelated"])
    ws2 = wb.create_sheet("cases")
    ws2.append(HEADER)
    ws2.append(["case_001", 1000, 0.3, 800])
    wb.save(p)

    ok = run({"file_path": str(p), "sheet_name": "cases"})
    assert ok["status"] == "success"
    assert len(ok["cases"]) == 1

    missing = run({"file_path": str(p), "sheet_name": "no_such_sheet"})
    assert missing["status"] == "failed"


def test_boolean_cells_are_row_errors_not_numbers(tmp_path: Path) -> None:
    """Codex 治理审 P2-1：Excel 布尔单元格（TRUE/FALSE）不是合法数值——
    float(True)==1.0 会静默吞入，必须判行级错误进 errors 不进 cases。"""
    p = _write_xlsx(tmp_path / "bool.xlsx", HEADER, [
        ["case_001", True, 0.3, 800],     # altitude_m 布尔
        ["case_002", 1000, False, 800],   # mach 布尔
        ["case_003", 1000, 0.3, 800],     # 合法对照
    ])
    result = run({"file_path": str(p)})
    assert result["status"] == "success"
    assert [c["case_id"] for c in result["cases"]] == ["case_003"]
    assert len(result["errors"]) == 2
    assert all("布尔值不是合法数值" in e["error"] for e in result["errors"])


def test_row_count_over_default_limit_rejected(tmp_path: Path) -> None:
    """P2 行数上限：1001 数据行（默认上限 1000）→ 整表诚实拒绝，绝不静默截断。"""
    rows = [[f"case_{i:04d}", 1000, 0.3, 800] for i in range(1, 1002)]
    p = _write_xlsx(tmp_path / "big.xlsx", HEADER, rows)
    result = run({"file_path": str(p)})
    assert result["status"] == "failed"
    assert "行数超上限" in result["error_message"] or "超上限" in result["error_message"]
    assert "分表" in result["error_message"]


def test_row_count_within_custom_max_rows_accepted(tmp_path: Path) -> None:
    """同一 1001 行表在 max_rows=2000 下正常解析（上限可放宽，行为对称）。"""
    rows = [[f"case_{i:04d}", 1000, 0.3, 800] for i in range(1, 1002)]
    p = _write_xlsx(tmp_path / "big_ok.xlsx", HEADER, rows)
    result = run({"file_path": str(p), "max_rows": 2000})
    assert result["status"] == "success"
    assert len(result["cases"]) == 1001


def test_deterministic_same_input_same_output(tmp_path: Path) -> None:
    p = _write_xlsx(tmp_path / "det.xlsx", HEADER, [["case_001", 1000, 0.3, 800]])
    assert run({"file_path": str(p)}) == run({"file_path": str(p)})
