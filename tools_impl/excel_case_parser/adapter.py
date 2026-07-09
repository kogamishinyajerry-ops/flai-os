"""excel_case_parser 适配器（mock=false，真实解析上传的 .xlsx case 表）。

契约（tool.yaml）：
- 首行 = 表头；必需列 case_id / altitude_m / mach / power_kw；可选列 bleed_flow_kgps；
  未声明的多余列一律忽略（README 注明）。
- 行级问题（缺值/非数值/超合理范围/重复 case_id）**不抛异常**：记入 errors
  并跳过该行——单行脏数据不摧毁整表解析（docs/05 §5 单 case 失败原则的解析端对应）。
- 解析器自身失败（文件不存在/非 xlsx/缺必需列/零有效行）才 status=failed。
- 适配器绝不抛裸异常（docs/03）：一切失败折叠为 {"status":"failed", "error_message":...}。

合理范围（比 mock 包线更宽——包线判定是 performance_disk_mock 的职责，
解析器只拦物理上不可能/明显笔误的值，绝不越权提前吞掉包线失败路径）：
  altitude_m ∈ [-500, 30000] / mach ∈ [0, 3] / power_kw ∈ (0, 20000] /
  bleed_flow_kgps ∈ [0, 50]
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

REQUIRED_COLUMNS = ("case_id", "altitude_m", "mach", "power_kw")
OPTIONAL_COLUMNS = ("bleed_flow_kgps",)

# (min, max, min_exclusive)
_RANGES: dict[str, tuple[float, float, bool]] = {
    "altitude_m": (-500.0, 30000.0, False),
    "mach": (0.0, 3.0, False),
    "power_kw": (0.0, 20000.0, True),
    "bleed_flow_kgps": (0.0, 50.0, False),
}


def _check_range(name: str, value: float) -> str | None:
    lo, hi, lo_exclusive = _RANGES[name]
    if lo_exclusive:
        if not (lo < value <= hi):
            return f"{name}={value} 超出合理范围（({lo}, {hi}]）"
    else:
        if not (lo <= value <= hi):
            return f"{name}={value} 超出合理范围（[{lo}, {hi}]）"
    return None


_DEFAULT_MAX_ROWS = 1000


def run(payload: dict[str, Any], context: dict[str, Any] | None = None) -> dict[str, Any]:
    del context  # V0.1 未用；保留形参符合 Tool Registry 调用约定
    file_path = Path(payload["file_path"])
    sheet_name = payload.get("sheet_name")
    max_rows = int(payload.get("max_rows", _DEFAULT_MAX_ROWS))

    if not file_path.is_file():
        return {"status": "failed", "error_message": f"文件不存在：{file_path}"}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - 契约要求绝不裸抛
        return {"status": "failed", "error_message": f"无法作为 xlsx 打开：{exc.__class__.__name__}: {exc}"}

    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                return {"status": "failed", "error_message": f"sheet 不存在：{sheet_name}（现有：{wb.sheetnames}）"}
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]

        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return {"status": "failed", "error_message": "空表：没有表头行"}

        header = [str(h).strip() if h is not None else "" for h in header_row]
        col_index: dict[str, int] = {}
        for name in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
            if name in header:
                col_index[name] = header.index(name)
        missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
        if missing:
            return {"status": "failed", "error_message": f"缺少必需列：{', '.join(missing)}（表头实为 {header}）"}

        param_columns = [c for c in ("altitude_m", "mach", "power_kw", "bleed_flow_kgps") if c in col_index]

        cases: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        seen_case_ids: set[str] = set()
        data_row_count = 0

        for excel_row, values in enumerate(rows, start=2):
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue  # 整行全空：静默跳过（不是数据行）

            data_row_count += 1
            if data_row_count > max_rows:
                # 超上限=解析器自身失败，诚实拒绝整表——绝不静默截断冒充全量结果。
                return {
                    "status": "failed",
                    "error_message": (
                        f"数据行数超上限（>{max_rows} 行），请分表后重新提交"
                        "（上限可经 max_rows 参数放宽，最大 5000）"
                    ),
                }

            def _cell(idx: int) -> Any:
                return values[idx] if idx < len(values) else None

            raw_case_id = _cell(col_index["case_id"])
            case_id = str(raw_case_id).strip() if raw_case_id is not None else ""
            if not case_id:
                errors.append({"row": excel_row, "case_id": None, "error": "case_id 缺失"})
                continue
            if case_id in seen_case_ids:
                errors.append({"row": excel_row, "case_id": case_id, "error": f"case_id 重复：{case_id}"})
                continue

            params: dict[str, float] = {}
            row_error: str | None = None
            for name in param_columns:
                raw = _cell(col_index[name])
                if raw is None or str(raw).strip() == "":
                    if name in OPTIONAL_COLUMNS:
                        continue  # 可选列缺值 = 不提供该参数
                    row_error = f"{name} 缺值"
                    break
                if isinstance(raw, bool):
                    # float(True)==1.0 会静默把 Excel 布尔单元格吞成数值（Codex 治理审 P2-1）
                    row_error = f"{name} 布尔值不是合法数值：{raw}"
                    break
                try:
                    value = float(raw)
                except (TypeError, ValueError):
                    row_error = f"{name} 非数值：{raw!r}"
                    break
                range_error = _check_range(name, value)
                if range_error is not None:
                    row_error = range_error
                    break
                params[name] = value

            if row_error is not None:
                errors.append({"row": excel_row, "case_id": case_id, "error": row_error})
                continue

            seen_case_ids.add(case_id)
            cases.append({"case_id": case_id, "params": params})

        if not cases:
            return {
                "status": "failed",
                "error_message": f"零有效 case 行（行级错误 {len(errors)} 条）",
            }
        return {"status": "success", "cases": cases, "errors": errors}
    except Exception as exc:  # noqa: BLE001 - 兜底：任何未预期解析异常折叠为 failed
        return {"status": "failed", "error_message": f"解析异常：{exc.__class__.__name__}: {exc}"}
    finally:
        wb.close()
