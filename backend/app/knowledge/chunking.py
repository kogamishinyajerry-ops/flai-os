"""多格式文档 → Chunk 语料（Knowledge 内核摄取层，ADR-0015）。

上游出处：COMAC_FDE core/ingest.py（收编复制适配，不 import 该仓）。与上游的刻意差异：
- csv 用 encoding="utf-8-sig"（真实语料常带 BOM，utf-8 会把 ﻿ 混进首列名）；
- docx 惰性 import，缺 python-docx 时抛 KnowledgeIngestError（可选依赖，诚实报错）；
- 全部失败路径统一 KnowledgeIngestError（fail-closed：未知后缀/PDF 一律 raise，
  绝不静默跳过——静默缺片会被误读为"语料里没有"）；
- ingest_dir 递归 rglob（上游为单层 iterdir），source=相对源目录的 POSIX 相对路径。

已知限制：doc_id 取文件 stem，同 stem 不同路径的文件允许共存，此时 chunk_id
（f"{doc_id}#{i}"）会跨文件碰撞不唯一；唯一定位以 source+fingerprint 为准。
"""

from __future__ import annotations

import csv
import errno
import hashlib
import io
import os
from dataclasses import dataclass
from pathlib import Path

from ..core.errors import KnowledgeIngestError

MAX_CHARS = 800


@dataclass(frozen=True)
class Chunk:
    """单个检索单元。source+fingerprint 是出处双钥，构造期即固化（frozen）。

    __post_init__ 强制出处双钥非空：frozen 只防已构造实例被改，非空校验把
    docs/06 §4「无出处禁止进入上下文」从调用链约定升级为类型层强制
    （loop-auditor Mode A Finding 3）。注意边界：这仍不能阻止拿着真格式
    假内容来构造——防的是漏填/空串，不是恶意伪造。
    """

    doc_id: str  # 源文件 stem
    chunk_id: str  # f"{doc_id}#{i}"，i 从 0 起（同 stem 文件间可碰撞，见模块 docstring）
    text: str
    source: str  # 相对 scope 源目录的 POSIX 相对路径（如 "manuals/em.md"）
    fingerprint: str  # sha256(文件字节)[:12]

    def __post_init__(self) -> None:
        if not (isinstance(self.source, str) and self.source.strip()):
            raise ValueError("Chunk.source 出处不得为空（docs/06 §4）")
        if not (isinstance(self.fingerprint, str) and self.fingerprint.strip()):
            raise ValueError("Chunk.fingerprint 出处指纹不得为空（docs/06 §4）")


# Windows 无 O_NOFOLLOW（位=0 退化普通打开，由 lstat 预检兜底——该平台创建
# symlink 需特权，威胁面显著小，诚实降级）。O_NOFOLLOW 拒开 symlink 的 errno
# 因平台而异：Linux/macOS=ELOOP，FreeBSD=EMLINK，NetBSD=EFTYPE。
_O_NOFOLLOW = getattr(os, "O_NOFOLLOW", 0)
_SYMLINK_ERRNOS = {
    e for e in (errno.ELOOP, getattr(errno, "EMLINK", None), getattr(errno, "EFTYPE", None))
    if e is not None
}


def _read_snapshot_no_follow(path: Path, source: str) -> bytes:
    """单次 no-follow 打开并整读字节快照（指纹与解析共用这一份，无第二次打开）。

    codex Wave1-R3 P1 + owner 裁决方案 b（2026-07-09）：语料摄取一律拒绝
    symlink——基于可变路径名的事后校验（resolve/stat 对照）存在本质 TOCTOU，
    open→resolve→stat 之间双换 symlink 即可绕过；收容必须由打开动作本身原子
    完成：O_NOFOLLOW 下最终组件是 symlink 直接拒开（内核判定，不存在
    检查/使用间隙）。域内 symlink 同罪：scope 语料目录由运维独占管理，
    symlink 无正当用例，"越界才拒"的区分本身就是竞态面。
    """
    if _O_NOFOLLOW == 0 and path.is_symlink() is True:
        raise KnowledgeIngestError(
            f"源文件 {source} 是 symlink（scope 源目录内不得有任何 symlink）"
        )
    try:
        fd = os.open(str(path), os.O_RDONLY | _O_NOFOLLOW)
    except OSError as exc:
        if exc.errno in _SYMLINK_ERRNOS:
            raise KnowledgeIngestError(
                f"源文件 {source} 是 symlink，已被 O_NOFOLLOW 原子拒开"
                "（scope 源目录内不得有任何 symlink）"
            ) from exc
        raise
    try:
        f = os.fdopen(fd, "rb")
    except Exception:
        os.close(fd)
        raise
    with f:
        return f.read()


def _reject_symlink_components(path: Path, root: Path, source: str) -> None:
    """root → 文件之间每个路径组件都不得是 symlink（方案 b 纵深防线）。

    O_NOFOLLOW 只约束最终组件；目录组件是 symlink 时整棵子树可指向根外。
    relative_to 用 lexical 比较不 resolve（resolve 会跟随 symlink，恰好抹掉
    待检对象）；不在根内直接拒。ingest_dir 的 rglob 本就不穿越目录 symlink，
    此检查覆盖直接调用 ingest_path 的路径与遍历后目录被替换的静态形态。
    """
    root = Path(root).absolute()
    try:
        rel_parts = Path(path).absolute().relative_to(root).parts
    except ValueError:
        raise KnowledgeIngestError(f"源文件 {source} 不在源目录 {root} 内") from None
    probe = root
    for part in rel_parts:
        probe = probe / part
        if probe.is_symlink() is True:
            raise KnowledgeIngestError(
                f"源文件 {source} 的路径组件 {part!r} 是 symlink"
                "（scope 源目录内不得有任何 symlink）"
            )


def _decode_text(raw: bytes, encoding: str) -> str:
    """bytes 快照 → str + 统一换行（\\r\\n、\\r → \\n）。

    等价快照化之前 text-mode（newline=None）打开的 universal newlines 语义
    （codex Wave1-R2 P1/P2）：\\r\\n 语料的段落边界（\\n\\n 切分）与 CR-only
    CSV 的解析行为都与旧实现一致——字节快照只改打开次数，不改解析语义。
    """
    return raw.decode(encoding).replace("\r\n", "\n").replace("\r", "\n")


def _merge(paras: list[str], doc_id: str, source: str, fp: str) -> list[Chunk]:
    """段落贪心合并：strip 后非空段依序拼接至 ≤MAX_CHARS，超限另起新 chunk。

    单段本身超 MAX_CHARS 先按 MAX_CHARS 硬切再进合并（codex Wave1-R1 P2）：
    否则超长段（超长 CSV 行/无空行长文）会整段成 chunk，击穿宣称的 800 字符
    上界，进而撑大检索命中与模型上下文。硬切可能落在词中间——检索按 jieba
    分词计分，边界词项损失可接受；出处双钥不受影响。
    """
    merged: list[str] = []
    buf = ""
    for p in paras:
        p = p.strip()
        if not p:
            continue
        for piece in (p[i : i + MAX_CHARS] for i in range(0, len(p), MAX_CHARS)):
            if buf and len(buf) + len(piece) + 1 > MAX_CHARS:
                merged.append(buf)
                buf = piece
            else:
                buf = f"{buf}\n{piece}" if buf else piece
    if buf:
        merged.append(buf)
    return [Chunk(doc_id, f"{doc_id}#{i}", t, source, fp) for i, t in enumerate(merged)]


def ingest_path(path: Path, *, source: str, containment_root: Path | None = None) -> list[Chunk]:
    """单文件 → list[Chunk]。source 由调用方显式传入（服务层给相对路径作出处）。

    空内容文件（全空白 txt、无行 xlsx）返回 []——空文件≠坏文件，空语料由
    service 层统一拒绝；格式不支持/可选依赖缺失则抛 KnowledgeIngestError。
    containment_root 非 None 时做 symlink 收容校验（ingest_dir 传源根目录）。
    """
    path = Path(path)
    suffix = path.suffix.lower()
    doc_id = path.stem
    # 指纹与解析绑定同一份字节快照（codex Wave1-R1 P2）：文件只打开一次，指纹
    # 从这份字节算、解析也从这份字节做——两次打开之间源文件被替换时，出处
    # 指纹会与正文脱钩。symlink 收容（codex Wave1-R3 P1，owner 裁决方案 b）：
    # 目录组件逐级预检 + 最终组件由 O_NOFOLLOW 在打开动作上原子拒绝。
    if containment_root is not None:
        _reject_symlink_components(path, containment_root, source)
    raw = _read_snapshot_no_follow(path, source)
    fp = hashlib.sha256(raw).hexdigest()[:12]
    if suffix in {".txt", ".md"}:
        paras = _decode_text(raw, "utf-8").split("\n\n")
    elif suffix == ".csv":
        # utf-8-sig：带 BOM 时剥掉 ﻿，无 BOM 时与 utf-8 等价（FDE retro 教训）。
        text = _decode_text(raw, "utf-8-sig")
        paras = ["; ".join(f"{k}={v}" for k, v in row.items()) for row in csv.DictReader(io.StringIO(text))]
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        ws = load_workbook(io.BytesIO(raw), read_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(c) for c in rows[0]]
        paras = ["; ".join(f"{h}={c}" for h, c in zip(header, r)) for r in rows[1:]]
    elif suffix == ".docx":
        try:
            import docx
        except ImportError as exc:
            raise KnowledgeIngestError(
                "python-docx 未安装，docx 解析不可用（可选依赖，离线环境见 README）"
            ) from exc
        paras = [p.text for p in docx.Document(io.BytesIO(raw)).paragraphs]
    elif suffix == ".pdf":
        raise KnowledgeIngestError("PDF 解析未接入（待内网侦察，诚实拒绝不静默跳过）")
    else:
        raise KnowledgeIngestError(
            f"不支持的格式: {suffix}（scope 源目录必须只含受支持格式，fail-closed 不静默跳过）"
        )
    return _merge(paras, doc_id, source, fp)


def ingest_dir(dir_path: Path) -> list[Chunk]:
    """目录（递归）→ list[Chunk]。

    - rglob 全量遍历，跳过任一路径分量以点开头的文件/目录（.hidden、.git/ 等）；
    - 按相对 POSIX 路径字符串排序，保证跨平台稳定顺序（索引缓存 manifest 依赖此序）；
    - source = 相对 dir_path 的 POSIX 路径（Windows 反斜杠不进出处字段）。
    """
    dir_path = Path(dir_path)
    rel_paths: list[Path] = []
    for p in dir_path.rglob("*"):
        if not p.is_file():
            continue
        rel = p.relative_to(dir_path)
        if any(part.startswith(".") for part in rel.parts):
            continue
        rel_paths.append(rel)
    # symlink 收容（codex Wave1-R1 P1 → R3 P1 终态，owner 裁决方案 b）单点
    # 收在 ingest_path：目录组件逐级预检 + 最终组件 O_NOFOLLOW 原子拒开。
    # 收集期不做预检——任何"先检后读"的路径名校验都有检查/使用间隙。
    # is_file() 会跟随符号链接，scope 内一个 `leak.md -> 仓外文件` 即可绕过
    # resolve_source_dir 的根校验，故 symlink 一律硬拒整次摄取（fail-closed，
    # 不静默跳过——静默跳过会把越界文件伪装成"语料里没有"）。
    chunks: list[Chunk] = []
    for rel in sorted(rel_paths, key=lambda r: r.as_posix()):
        chunks.extend(
            ingest_path(dir_path / rel, source=rel.as_posix(), containment_root=dir_path)
        )
    return chunks
