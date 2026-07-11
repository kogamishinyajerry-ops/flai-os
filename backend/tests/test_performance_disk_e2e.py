"""performance_disk_agent M3 端到端验收（真 create_app + JobRunner，全走 API）。

链路：API 上传 case_table_50.xlsx → POST /api/tasks（input_file_ids）→
JobRunner.run_once() → 断言任务终态/三产物/汇总表回读对账/samples.jsonl
逐行对账/事件时间轴（agent_log 折叠的 case_failed）/tool_runs mock=true 如实。

期望口径来自 eval_cases/case_001.json（50 行 = 48 有效 case = 45 成功 + 3
mock 包线失败，2 行解析行级错误）——xlsx 由 generate_case_table.py 确定性
生成，本测试若与口径不符即为链路回归而非数据漂移。

另覆盖：全失败表（全部 altitude>15000）任务仍 completed（批量语义，
ADR-0010）；空表（仅表头）任务 failed 诚实；无输入文件任务 failed。
"""

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Iterator

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.app.jobs.runner import JobRunner
from backend.app.storage import repos

REPO_ROOT = Path(__file__).resolve().parents[2]
CASE_TABLE = REPO_ROOT / "agents" / "performance_disk_agent" / "eval_cases" / "case_table_50.xlsx"

# eval_cases/case_001.json 的期望口径（单一事实源：两处都改才算改口径）
EXPECTED = json.loads(
    (REPO_ROOT / "agents" / "performance_disk_agent" / "eval_cases" / "case_001.json").read_text(encoding="utf-8")
)["expected"]


@pytest.fixture()
def client(app_env) -> Iterator[TestClient]:
    c, _ = app_env
    yield c


def _upload_xlsx(client: TestClient, content: bytes, filename: str = "case_table.xlsx") -> str:
    resp = client.post(
        "/api/files/upload",
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    return resp.json()["id"]


def _create_and_run(client: TestClient, app, file_ids: list[str]) -> dict:
    resp = client.post(
        "/api/tasks",
        json={
            "agent_id": "performance_disk_agent",
            "inputs": {},
            "input_file_ids": file_ids,
            "created_by": "m3_e2e",
        },
    )
    assert resp.status_code == 200
    task_id = resp.json()["id"]
    runner = JobRunner(app.state.runtime, app.state.conn_factory)
    assert runner.run_once() is True
    return client.get(f"/api/tasks/{task_id}").json()


def _outputs_by_name(client: TestClient, app, task: dict) -> dict[str, bytes]:
    """下载全部输出产物，按 filename 索引其字节内容。"""
    conn = app.state.conn_factory()
    try:
        records = [repos.get_file(conn, fid) for fid in task["output_file_ids"]]
    finally:
        conn.close()
    out: dict[str, bytes] = {}
    for record in records:
        assert record is not None
        resp = client.get(f"/api/files/{record['id']}/download")
        assert resp.status_code == 200
        out[record["filename"]] = resp.content
    return out


def _build_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["case_id", "altitude_m", "mach", "power_kw"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 主 E2E：50 行示例表全链路 ─────────────────────────────────────────────


def test_m3_e2e_full_chain_with_50_case_table(app_env) -> None:
    client, app = app_env
    file_id = _upload_xlsx(client, CASE_TABLE.read_bytes(), "case_table_50.xlsx")

    task = _create_and_run(client, app, [file_id])
    task_id = task["id"]

    # ① 任务 completed（批量语义：3 个失败 case 不摧毁任务）
    assert task["status"] == EXPECTED["task_status"] == "completed"

    # ② 三产物齐且可下载
    outputs = _outputs_by_name(client, app, task)
    assert set(outputs.keys()) == set(EXPECTED["artifacts"])

    # ③ result_summary.xlsx 回读：行数与 status 分布对账 + mock 水印（五落点之表格）
    wb = openpyxl.load_workbook(io.BytesIO(outputs["result_summary.xlsx"]))
    assert "声明" in wb.sheetnames, "表格产物必须带「声明」sheet（docs/03 §3 第五落点）"
    assert wb.sheetnames[0] == "result_summary", "数据 sheet 必须保持第一"
    notice_text = wb["声明"]["A1"].value
    assert notice_text and "无任何工程意义" in notice_text

    ws = wb["result_summary"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]
    assert len(data_rows) == EXPECTED["total_cases"]
    status_col = header.index("status")
    error_col = header.index("error_message")
    mock_col = header.index("mock")
    statuses = [r[status_col] for r in data_rows]
    assert statuses.count("success") == EXPECTED["ok_count"]
    assert statuses.count("failed") == EXPECTED["failed_count"]
    for r in data_rows:
        if r[status_col] == "failed":
            assert r[error_col] and "超出 mock 包线" in r[error_col]
    assert all(r[mock_col] is True for r in data_rows), "数据行 mock 列必须全 True（逐行如实标注）"

    # ④ samples.jsonl：行数=有效 case 数；失败行有 error_message；每行 mock 如实
    sample_lines = [json.loads(l) for l in outputs["samples.jsonl"].decode("utf-8").splitlines()]
    assert len(sample_lines) == EXPECTED["total_cases"]
    failed_lines = [s for s in sample_lines if s["status"] == "failed"]
    assert len(failed_lines) == EXPECTED["failed_count"]
    for s in failed_lines:
        assert s["error_message"], "失败样本行必须携带 error_message"
    assert all(s["mock"] is True for s in sample_lines), "样本必须逐行如实标注 mock=true（docs/03 §3）"

    # ⑤ task_report.md：mock 声明 + 计数如实
    report = outputs["task_report.md"].decode("utf-8")
    assert "MOCK 声明" in report and "无任何工程意义" in report
    assert f"计算成功：{EXPECTED['ok_count']}" in report
    assert f"计算失败：{EXPECTED['failed_count']}" in report

    # ⑥ 事件时间轴：agent_log 折叠保留 workflow_event_type=case_failed（3 次）
    events = client.get(f"/api/tasks/{task_id}/events").json()
    folded = [
        e for e in events
        if e["event_type"] == "agent_log" and e["payload"].get("workflow_event_type") == "case_failed"
    ]
    assert len(folded) == EXPECTED["failed_count"]
    assert all(f["payload"].get("mock") is True for f in folded)
    assert any(
        e["event_type"] == "agent_log" and e["payload"].get("workflow_event_type") == "summary_generated"
        for e in events
    )

    # ⑦ tool_runs：performance_disk_mock 每 case 一行，mock=true 如实入库
    conn = app.state.conn_factory()
    try:
        runs = repos.list_tool_runs(conn, task_id)
    finally:
        conn.close()
    mock_runs = [r for r in runs if r["tool_id"] == "performance_disk_mock"]
    assert len(mock_runs) == EXPECTED["total_cases"]
    assert all(r["mock"] is True for r in mock_runs)
    # 解析器与汇总写出器各一次，且 mock=false 如实
    assert sum(1 for r in runs if r["tool_id"] == "excel_case_parser") == 1
    assert sum(1 for r in runs if r["tool_id"] == "excel_summary_writer") == 1
    assert all(r["mock"] is False for r in runs if r["tool_id"] != "performance_disk_mock")


# ── 全失败表：任务仍 completed，汇总全 failed（批量语义，ADR-0010）─────────


def test_all_cases_fail_task_still_completed(app_env) -> None:
    client, app = app_env
    content = _build_xlsx([
        [f"case_{i:03d}", 16000 + i * 100, 0.5, 1000] for i in range(1, 6)
    ])
    file_id = _upload_xlsx(client, content, "all_fail.xlsx")

    task = _create_and_run(client, app, [file_id])
    assert task["status"] == "completed", "全失败 case 仍是任务成功（批量作业语义）"

    outputs = _outputs_by_name(client, app, task)
    ws = openpyxl.load_workbook(io.BytesIO(outputs["result_summary.xlsx"]))["result_summary"]
    rows = list(ws.iter_rows(values_only=True))
    status_col = rows[0].index("status")
    statuses = [r[status_col] for r in rows[1:]]
    assert len(statuses) == 5
    assert statuses.count("failed") == 5
    assert statuses.count("success") == 0


# ── 多附件（Codex P2-2：绝不盲取 files[0]）───────────────────────────────


def test_txt_plus_xlsx_uses_the_xlsx(app_env) -> None:
    """txt 在前 + xlsx 在后：必须按后缀选中 xlsx（旧实现盲取 files[0] 会拿 txt 炸解析）。"""
    client, app = app_env
    txt_resp = client.post(
        "/api/files/upload",
        files={"file": ("notes.txt", "这是一段无关说明".encode("utf-8"), "text/plain")},
    )
    assert txt_resp.status_code == 200
    txt_id = txt_resp.json()["id"]

    xlsx_id = _upload_xlsx(client, _build_xlsx([
        ["case_001", 1000, 0.3, 800],
        ["case_002", 2000, 0.4, 900],
    ]), "small.xlsx")

    # txt 在前：input_file_ids 顺序 = [txt, xlsx]
    task = _create_and_run(client, app, [txt_id, xlsx_id])
    assert task["status"] == "completed"

    outputs = _outputs_by_name(client, app, task)
    ws = openpyxl.load_workbook(io.BytesIO(outputs["result_summary.xlsx"]))["result_summary"]
    data_rows = list(ws.iter_rows(values_only=True))[1:]
    assert len(data_rows) == 2, "必须解析的是 xlsx（2 case），不是 txt"


def test_two_xlsx_files_task_failed_honestly(app_env) -> None:
    """双 xlsx：拒绝猜测用户想算哪张表，任务 failed 且信息写明数量。"""
    client, app = app_env
    id_a = _upload_xlsx(client, _build_xlsx([["case_001", 1000, 0.3, 800]]), "a.xlsx")
    id_b = _upload_xlsx(client, _build_xlsx([["case_001", 1000, 0.3, 800]]), "b.xlsx")

    task = _create_and_run(client, app, [id_a, id_b])
    assert task["status"] == "failed"
    assert "2 个 xlsx" in task["error_message"]
    assert "只上传一个" in task["error_message"]


# ── 空表 / 无输入文件：任务 failed 诚实 ───────────────────────────────────


def test_header_only_table_task_failed_honestly(app_env) -> None:
    client, app = app_env
    file_id = _upload_xlsx(client, _build_xlsx([]), "empty.xlsx")

    task = _create_and_run(client, app, [file_id])
    assert task["status"] == "failed"
    assert "解析失败" in task["error_message"]
    assert task["output_file_ids"] == [], "失败任务不应注册产物"


def test_no_input_file_task_failed_honestly(app_env) -> None:
    client, app = app_env
    task = _create_and_run(client, app, [])
    assert task["status"] == "failed"
    assert "无输入文件" in task["error_message"]
